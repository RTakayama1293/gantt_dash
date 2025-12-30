"""
EEZO 2026 タスクダッシュボード
============================

北海道食材EC（EEZO）の年間タスクを可視化するガントチャートダッシュボード。

使用方法:
    python src/app.py

ブラウザでアクセス:
    http://127.0.0.1:8050
"""

import pandas as pd
from dash import Dash, html, dcc, callback, Output, Input, State, no_update, ctx
import dash_bootstrap_components as dbc
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date, timedelta
from io import BytesIO
import os
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 定数定義
# =============================================================================

DATA_PATH = "data/raw/eezo_2026_weekly_tasks.csv"
OUTPUT_DIR = "output"

# カテゴリ色設定（CLAUDE.md準拠）
CATEGORY_COLORS: dict[str, str] = {
    "プラットフォーム実装": "#3498db",  # 青
    "UX動線": "#9b59b6",                # 紫
    "商品コンテンツ": "#e74c3c",        # 赤
    "集客販促": "#f39c12",              # オレンジ
    "データ活用": "#2ecc71",            # 緑
}

# カテゴリのデフォルト順序
DEFAULT_CATEGORY_ORDER: list[str] = [
    "プラットフォーム実装",
    "UX動線",
    "商品コンテンツ",
    "集客販促",
    "データ活用",
]

# 担当者色設定
ASSIGNEE_COLORS: dict[str, str] = {
    "松永": "#2980b9",  # 青系
    "高山": "#c0392b",  # 赤系
}

# 担当者のデフォルト順序
DEFAULT_ASSIGNEE_ORDER: list[str] = ["松永", "高山"]


# =============================================================================
# データ読み込み
# =============================================================================

def load_data(filepath: str) -> pd.DataFrame:
    """
    CSVファイルを読み込み、日付をパースする。

    Args:
        filepath: CSVファイルのパス

    Returns:
        pd.DataFrame: 読み込んだデータフレーム
    """
    df = pd.read_csv(filepath, encoding="utf-8")
    df["開始日"] = pd.to_datetime(df["開始日"], format="%Y/%m/%d")
    df["終了日"] = pd.to_datetime(df["終了日"], format="%Y/%m/%d")

    # マイルストーンフラグを追加（★マーク付き）
    df["is_milestone"] = df["成果物/マイルストーン"].str.contains("★", na=False)

    # タスクIDを追加
    df["task_id"] = range(1, len(df) + 1)

    # 期間（日数）を計算
    df["期間"] = (df["終了日"] - df["開始日"]).dt.days + 1

    return df


def sort_dataframe(
    df: pd.DataFrame,
    sort_by: str,
    sort_order: str,
    category_order: list[str],
    assignee_order: list[str]
) -> pd.DataFrame:
    """
    データフレームをソートする。

    Args:
        df: ソート対象のデータフレーム
        sort_by: ソートキー（"開始日", "担当者", "カテゴリ"）
        sort_order: ソート順（"asc", "desc"）
        category_order: カテゴリの並び順
        assignee_order: 担当者の並び順

    Returns:
        pd.DataFrame: ソート済みデータフレーム
    """
    ascending = sort_order == "asc"
    result_df = df.copy()

    if sort_by == "開始日":
        result_df = result_df.sort_values(
            ["開始日", "担当者"],
            ascending=[ascending, True]
        )
    elif sort_by == "担当者":
        # カスタム順序でソート
        result_df["担当者_order"] = result_df["担当者"].apply(
            lambda x: assignee_order.index(x) if x in assignee_order else 999
        )
        result_df = result_df.sort_values(
            ["担当者_order", "開始日"],
            ascending=[ascending, True]
        )
        result_df = result_df.drop(columns=["担当者_order"])
    elif sort_by == "カテゴリ":
        # カスタム順序でソート
        result_df["カテゴリ_order"] = result_df["カテゴリ"].apply(
            lambda x: category_order.index(x) if x in category_order else 999
        )
        result_df = result_df.sort_values(
            ["カテゴリ_order", "開始日"],
            ascending=[ascending, True]
        )
        result_df = result_df.drop(columns=["カテゴリ_order"])

    return result_df


# =============================================================================
# ガントチャート生成
# =============================================================================

def create_gantt_chart(
    df: pd.DataFrame,
    color_by: str = "カテゴリ",
    granularity: str = "week",
    group_by: str = "none",
    show_today_line: bool = True
) -> go.Figure:
    """
    ガントチャートを生成する。

    Args:
        df: タスクデータフレーム
        color_by: 色分けの基準（"カテゴリ" or "担当者"）
        granularity: 時間粒度（"day", "week", "month"）
        group_by: グループ化（"none", "担当者", "カテゴリ"）
        show_today_line: 今日線を表示するか

    Returns:
        go.Figure: Plotlyのガントチャート
    """
    if df.empty:
        fig = go.Figure()
        fig.update_layout(
            title="データがありません",
            height=200,
            paper_bgcolor="white",
            plot_bgcolor="white"
        )
        return fig

    color_map = CATEGORY_COLORS if color_by == "カテゴリ" else ASSIGNEE_COLORS

    # グループ化に応じてY軸のラベルを調整
    df_chart = df.copy()
    if group_by == "担当者":
        df_chart["y_label"] = df_chart["担当者"] + " | " + df_chart["タスク"]
    elif group_by == "カテゴリ":
        df_chart["y_label"] = df_chart["カテゴリ"] + " | " + df_chart["タスク"]
    else:
        df_chart["y_label"] = df_chart["タスク"]

    # ガントチャート作成
    fig = px.timeline(
        df_chart,
        x_start="開始日",
        x_end="終了日",
        y="y_label",
        color=color_by,
        color_discrete_map=color_map,
        custom_data=["四半期", "週番号", "担当者", "カテゴリ", "成果物/マイルストーン", "期間"],
    )

    # ホバーテンプレート設定
    fig.update_traces(
        hovertemplate=(
            "<b>%{y}</b><br>"
            "期間: %{x|%Y/%m/%d} 〜 %{customdata[5]}日間<br>"
            "四半期: %{customdata[0]} / %{customdata[1]}<br>"
            "担当者: %{customdata[2]}<br>"
            "カテゴリ: %{customdata[3]}<br>"
            "成果物: %{customdata[4]}<br>"
            "<extra></extra>"
        )
    )

    # マイルストーン強調（★付きタスク）
    milestones = df_chart[df_chart["is_milestone"]]
    if not milestones.empty:
        fig.add_trace(go.Scatter(
            x=milestones["終了日"],
            y=milestones["y_label"],
            mode="markers",
            marker=dict(
                symbol="star",
                size=14,
                color="gold",
                line=dict(color="black", width=1)
            ),
            name="★マイルストーン",
            hoverinfo="skip"
        ))

    # 今日線を追加
    if show_today_line:
        today = datetime.now()
        chart_min_date = df_chart["開始日"].min()
        chart_max_date = df_chart["終了日"].max()

        if chart_min_date <= today <= chart_max_date:
            fig.add_vline(
                x=today,
                line_dash="dash",
                line_color="red",
                line_width=2,
                annotation_text="今日",
                annotation_position="top"
            )

    # レイアウト調整
    chart_height = max(500, len(df_chart) * 28)
    fig.update_layout(
        height=chart_height,
        xaxis_title="日付",
        yaxis_title="",
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor="rgba(255,255,255,0.8)"
        ),
        margin=dict(l=300, r=50, t=80, b=50),
        paper_bgcolor="white",
        plot_bgcolor="#fafafa",
        yaxis=dict(
            categoryorder="array",
            categoryarray=df_chart["y_label"].tolist()[::-1]
        ),
        font=dict(family="Noto Sans JP, sans-serif"),
    )

    # X軸の粒度設定
    if granularity == "day":
        fig.update_xaxes(dtick="D1", tickformat="%m/%d", tickangle=45)
    elif granularity == "week":
        fig.update_xaxes(dtick="D7", tickformat="%m/%d")
    else:  # month
        fig.update_xaxes(dtick="M1", tickformat="%Y/%m")

    # グリッド線
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor="rgba(0,0,0,0.1)")
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor="rgba(0,0,0,0.05)")

    return fig


# =============================================================================
# Excelガントチャート出力
# =============================================================================

def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    """16進数カラーコードをRGBタプルに変換"""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def create_excel_gantt_chart(
    df: pd.DataFrame,
    granularity: str,
    color_by: str
) -> BytesIO:
    """
    Excelガントチャートを生成する。

    Args:
        df: タスクデータフレーム（ソート済み）
        granularity: 時間粒度（"day", "week", "month"）
        color_by: 色分けの基準（"カテゴリ" or "担当者"）

    Returns:
        BytesIO: Excelファイルのバイトストリーム
    """
    if df.empty:
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "ガントチャート"
        ws["A1"] = "データがありません"
        wb.save(output)
        output.seek(0)
        return output

    wb = Workbook()
    ws = wb.active
    ws.title = "ガントチャート"

    # 色マップの選択
    color_map = CATEGORY_COLORS if color_by == "カテゴリ" else ASSIGNEE_COLORS

    # 日付範囲の計算
    start_date = df["開始日"].min()
    end_date = df["終了日"].max()

    # 粒度に応じて日付リストを生成
    if granularity == "day":
        date_list = pd.date_range(start=start_date, end=end_date, freq="D")
        date_format = "%m/%d"
        header_format = "%m/%d\n(%a)"
    elif granularity == "week":
        # 週の開始日（月曜日）に揃える
        week_start = start_date - timedelta(days=start_date.weekday())
        date_list = pd.date_range(start=week_start, end=end_date, freq="W-MON")
        date_format = "%m/%d"
        header_format = "%m/%d"
    else:  # month
        # 月初に揃える
        month_start = start_date.replace(day=1)
        date_list = pd.date_range(start=month_start, end=end_date, freq="MS")
        date_format = "%Y/%m"
        header_format = "%Y/%m"

    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    task_font = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    # ヘッダー行1: 年/月（月粒度の場合は四半期）
    # ヘッダー行2: 日付
    header_row = 2
    task_start_row = 3

    # 列の設定
    # A列: No.
    # B列: 担当者
    # C列: カテゴリ
    # D列: タスク名
    # E列以降: 日付

    # ヘッダー設定
    ws.cell(row=1, column=1, value="No.").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border
    ws.column_dimensions["A"].width = 5

    ws.cell(row=1, column=2, value="担当者").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = center_align
    ws.cell(row=1, column=2).border = thin_border
    ws.column_dimensions["B"].width = 8

    ws.cell(row=1, column=3, value="カテゴリ").font = header_font
    ws.cell(row=1, column=3).fill = header_fill
    ws.cell(row=1, column=3).alignment = center_align
    ws.cell(row=1, column=3).border = thin_border
    ws.column_dimensions["C"].width = 18

    ws.cell(row=1, column=4, value="タスク").font = header_font
    ws.cell(row=1, column=4).fill = header_fill
    ws.cell(row=1, column=4).alignment = center_align
    ws.cell(row=1, column=4).border = thin_border
    ws.column_dimensions["D"].width = 45

    # 日付ヘッダー
    date_start_col = 5
    for i, d in enumerate(date_list):
        col = date_start_col + i
        cell = ws.cell(row=1, column=col, value=d.strftime(header_format))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # 列幅設定
        if granularity == "day":
            ws.column_dimensions[get_column_letter(col)].width = 5
        elif granularity == "week":
            ws.column_dimensions[get_column_letter(col)].width = 6
        else:
            ws.column_dimensions[get_column_letter(col)].width = 8

    # タスク行の出力
    for row_idx, (_, task) in enumerate(df.iterrows(), start=2):
        # No.
        ws.cell(row=row_idx, column=1, value=row_idx - 1).font = task_font
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=1).border = thin_border

        # 担当者
        ws.cell(row=row_idx, column=2, value=task["担当者"]).font = task_font
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=2).border = thin_border

        # カテゴリ
        ws.cell(row=row_idx, column=3, value=task["カテゴリ"]).font = task_font
        ws.cell(row=row_idx, column=3).alignment = left_align
        ws.cell(row=row_idx, column=3).border = thin_border

        # タスク名（マイルストーンは★を付ける）
        task_name = task["タスク"]
        if task["is_milestone"]:
            task_name = "★ " + task_name
        ws.cell(row=row_idx, column=4, value=task_name).font = task_font
        ws.cell(row=row_idx, column=4).alignment = left_align
        ws.cell(row=row_idx, column=4).border = thin_border

        # 色の取得
        color_key = task[color_by]
        hex_color = color_map.get(color_key, "#999999")
        rgb = hex_to_rgb(hex_color)
        fill_color = PatternFill(
            start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
            end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
            fill_type="solid"
        )

        # 日付セルの塗りつぶし
        task_start = task["開始日"]
        task_end = task["終了日"]

        for i, d in enumerate(date_list):
            col = date_start_col + i
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border

            # 粒度に応じた期間判定
            if granularity == "day":
                period_start = d
                period_end = d
            elif granularity == "week":
                period_start = d
                period_end = d + timedelta(days=6)
            else:  # month
                period_start = d
                # 月末を計算
                if d.month == 12:
                    period_end = d.replace(year=d.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    period_end = d.replace(month=d.month + 1, day=1) - timedelta(days=1)

            # タスク期間と重なるかチェック
            if task_start <= period_end and task_end >= period_start:
                cell.fill = fill_color

    # 行の高さ設定
    for row in range(1, len(df) + 2):
        ws.row_dimensions[row].height = 20

    # ウィンドウ枠の固定（ヘッダーとタスク名列）
    ws.freeze_panes = "E2"

    # 凡例シートの追加
    ws_legend = wb.create_sheet(title="凡例")

    ws_legend.cell(row=1, column=1, value="【色の凡例】").font = Font(bold=True, size=11)

    legend_items = CATEGORY_COLORS if color_by == "カテゴリ" else ASSIGNEE_COLORS
    for i, (name, hex_color) in enumerate(legend_items.items(), start=3):
        rgb = hex_to_rgb(hex_color)
        fill = PatternFill(
            start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
            end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
            fill_type="solid"
        )
        ws_legend.cell(row=i, column=1).fill = fill
        ws_legend.cell(row=i, column=1).border = thin_border
        ws_legend.cell(row=i, column=2, value=name).font = Font(size=10)
        ws_legend.column_dimensions["A"].width = 5
        ws_legend.column_dimensions["B"].width = 20

    # サマリー情報
    summary_row = len(legend_items) + 5
    ws_legend.cell(row=summary_row, column=1, value="【サマリー】").font = Font(bold=True, size=11)
    ws_legend.cell(row=summary_row + 1, column=1, value="総タスク数:")
    ws_legend.cell(row=summary_row + 1, column=2, value=len(df))
    ws_legend.cell(row=summary_row + 2, column=1, value="期間:")
    ws_legend.cell(row=summary_row + 2, column=2, value=f"{start_date.strftime('%Y/%m/%d')} 〜 {end_date.strftime('%Y/%m/%d')}")
    ws_legend.cell(row=summary_row + 3, column=1, value="マイルストーン数:")
    ws_legend.cell(row=summary_row + 3, column=2, value=df["is_milestone"].sum())

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# Dashアプリケーション
# =============================================================================

# アプリ初期化
app = Dash(
    __name__,
    external_stylesheets=[
        dbc.themes.FLATLY,
        "https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;700&display=swap"
    ],
    suppress_callback_exceptions=True
)
app.title = "EEZO 2026 タスクダッシュボード"

# データ読み込み
df = load_data(DATA_PATH)

# 日付範囲の計算
min_date = df["開始日"].min()
max_date = df["終了日"].max()
date_range_days = (max_date - min_date).days


# =============================================================================
# レイアウト
# =============================================================================

# ヘッダー
header = dbc.Navbar(
    dbc.Container([
        dbc.Row([
            dbc.Col([
                html.H3("EEZO 2026 タスクダッシュボード", className="text-white mb-0"),
                html.Small("北海道食材EC 年間プロジェクト管理", className="text-light")
            ]),
        ], align="center", className="flex-grow-1"),
        dbc.Row([
            dbc.Col([
                dbc.Button(
                    [html.I(className="fas fa-file-excel me-2"), "Excel出力"],
                    id="download-btn",
                    color="success",
                    className="me-2"
                ),
                dcc.Download(id="download-excel")
            ])
        ], align="center"),
    ], fluid=True),
    color="primary",
    dark=True,
    className="mb-3"
)

# フィルターカード
filter_card = dbc.Card([
    dbc.CardHeader([
        html.I(className="fas fa-filter me-2"),
        "フィルター・表示設定"
    ], className="fw-bold"),
    dbc.CardBody([
        # 第1行: フィルター
        dbc.Row([
            dbc.Col([
                dbc.Label("四半期", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="quarter-filter",
                    options=[{"label": q, "value": q} for q in sorted(df["四半期"].unique())],
                    value=sorted(df["四半期"].unique().tolist()),
                    multi=True,
                    placeholder="四半期を選択..."
                )
            ], md=3),
            dbc.Col([
                dbc.Label("担当者", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="assignee-filter",
                    options=[{"label": a, "value": a} for a in df["担当者"].unique()],
                    value=df["担当者"].unique().tolist(),
                    multi=True,
                    placeholder="担当者を選択..."
                )
            ], md=3),
            dbc.Col([
                dbc.Label("カテゴリ", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="category-filter",
                    options=[{"label": c, "value": c} for c in df["カテゴリ"].unique()],
                    value=df["カテゴリ"].unique().tolist(),
                    multi=True,
                    placeholder="カテゴリを選択..."
                )
            ], md=6),
        ], className="mb-3"),

        # 第2行: 表示設定
        dbc.Row([
            dbc.Col([
                dbc.Label("期間粒度", className="fw-bold text-muted small"),
                dbc.RadioItems(
                    id="granularity",
                    options=[
                        {"label": "日", "value": "day"},
                        {"label": "週", "value": "week"},
                        {"label": "月", "value": "month"}
                    ],
                    value="week",
                    inline=True,
                    className="mt-1"
                )
            ], md=2),
            dbc.Col([
                dbc.Label("グループ化", className="fw-bold text-muted small"),
                dbc.RadioItems(
                    id="group-by",
                    options=[
                        {"label": "なし", "value": "none"},
                        {"label": "担当者別", "value": "担当者"},
                        {"label": "カテゴリ別", "value": "カテゴリ"}
                    ],
                    value="none",
                    inline=True,
                    className="mt-1"
                )
            ], md=3),
            dbc.Col([
                dbc.Label("色分け", className="fw-bold text-muted small"),
                dbc.RadioItems(
                    id="color-by",
                    options=[
                        {"label": "カテゴリ", "value": "カテゴリ"},
                        {"label": "担当者", "value": "担当者"}
                    ],
                    value="カテゴリ",
                    inline=True,
                    className="mt-1"
                )
            ], md=2),
            dbc.Col([
                dbc.Label("今日線", className="fw-bold text-muted small"),
                dbc.Checklist(
                    id="show-today-line",
                    options=[{"label": "表示", "value": True}],
                    value=[True],
                    inline=True,
                    className="mt-1"
                )
            ], md=2),
        ], className="mb-3"),

        # 第3行: ソート設定（強化）
        dbc.Row([
            dbc.Col([
                dbc.Label("ソート項目", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="sort-by",
                    options=[
                        {"label": "開始日", "value": "開始日"},
                        {"label": "担当者", "value": "担当者"},
                        {"label": "カテゴリ", "value": "カテゴリ"}
                    ],
                    value="開始日",
                    clearable=False,
                    className="mt-1"
                )
            ], md=2),
            dbc.Col([
                dbc.Label("昇順/降順", className="fw-bold text-muted small"),
                dbc.RadioItems(
                    id="sort-order",
                    options=[
                        {"label": "昇順 ↑", "value": "asc"},
                        {"label": "降順 ↓", "value": "desc"}
                    ],
                    value="asc",
                    inline=True,
                    className="mt-1"
                )
            ], md=2),
            dbc.Col([
                dbc.Label("カテゴリ並び順", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="category-order",
                    options=[{"label": c, "value": c} for c in DEFAULT_CATEGORY_ORDER],
                    value=DEFAULT_CATEGORY_ORDER,
                    multi=True,
                    placeholder="ドラッグで並べ替え..."
                )
            ], md=4),
            dbc.Col([
                dbc.Label("担当者並び順", className="fw-bold text-muted small"),
                dcc.Dropdown(
                    id="assignee-order",
                    options=[{"label": a, "value": a} for a in DEFAULT_ASSIGNEE_ORDER],
                    value=DEFAULT_ASSIGNEE_ORDER,
                    multi=True,
                    placeholder="ドラッグで並べ替え..."
                )
            ], md=2),
        ], className="mb-3"),

        # 第4行: 日付範囲スライダー
        dbc.Row([
            dbc.Col([
                dbc.Label("日付範囲", className="fw-bold text-muted small"),
                dcc.RangeSlider(
                    id="date-range-slider",
                    min=0,
                    max=date_range_days,
                    step=7,
                    value=[0, date_range_days],
                    marks={
                        0: min_date.strftime("%Y/%m"),
                        date_range_days // 4: (min_date + pd.Timedelta(days=date_range_days // 4)).strftime("%Y/%m"),
                        date_range_days // 2: (min_date + pd.Timedelta(days=date_range_days // 2)).strftime("%Y/%m"),
                        date_range_days * 3 // 4: (min_date + pd.Timedelta(days=date_range_days * 3 // 4)).strftime("%Y/%m"),
                        date_range_days: max_date.strftime("%Y/%m"),
                    },
                    tooltip={"placement": "bottom", "always_visible": False}
                )
            ], md=12),
        ]),
    ])
], className="mb-3 shadow-sm")

# サマリーカード
summary_card = dbc.Card([
    dbc.CardBody([
        dbc.Row([
            dbc.Col([
                html.Div([
                    html.Span("📊 ", style={"fontSize": "1.5em"}),
                    html.Span("タスク数", className="text-muted small d-block"),
                    html.Span(id="total-tasks", className="h4 fw-bold text-primary")
                ], className="text-center")
            ], md=2),
            dbc.Col([
                html.Div([
                    html.Span("📅 ", style={"fontSize": "1.5em"}),
                    html.Span("期間", className="text-muted small d-block"),
                    html.Span(id="date-range", className="h6")
                ], className="text-center")
            ], md=3),
            dbc.Col([
                html.Div([
                    html.Span("👥 ", style={"fontSize": "1.5em"}),
                    html.Span("担当者別", className="text-muted small d-block"),
                    html.Div(id="assignee-summary")
                ], className="text-center")
            ], md=3),
            dbc.Col([
                html.Div([
                    html.Span("📁 ", style={"fontSize": "1.5em"}),
                    html.Span("カテゴリ別", className="text-muted small d-block"),
                    html.Div(id="category-summary")
                ], className="text-center")
            ], md=4),
        ], align="center")
    ])
], className="mb-3 shadow-sm")

# メインレイアウト
app.layout = dbc.Container([
    header,
    filter_card,
    summary_card,
    dbc.Card([
        dbc.CardBody([
            dcc.Loading(
                dcc.Graph(
                    id="gantt-chart",
                    config={
                        "displayModeBar": True,
                        "displaylogo": False,
                        "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                        "toImageButtonOptions": {
                            "format": "png",
                            "filename": "eezo_gantt_chart",
                            "height": 1200,
                            "width": 1800,
                            "scale": 2
                        }
                    }
                ),
                type="circle",
                color="#3498db"
            )
        ])
    ], className="shadow-sm")
], fluid=True, className="pb-4")


# =============================================================================
# コールバック
# =============================================================================

@callback(
    [Output("gantt-chart", "figure"),
     Output("total-tasks", "children"),
     Output("date-range", "children"),
     Output("assignee-summary", "children"),
     Output("category-summary", "children")],
    [Input("quarter-filter", "value"),
     Input("assignee-filter", "value"),
     Input("category-filter", "value"),
     Input("color-by", "value"),
     Input("granularity", "value"),
     Input("group-by", "value"),
     Input("sort-by", "value"),
     Input("sort-order", "value"),
     Input("category-order", "value"),
     Input("assignee-order", "value"),
     Input("date-range-slider", "value"),
     Input("show-today-line", "value")]
)
def update_dashboard(
    quarters: list[str],
    assignees: list[str],
    categories: list[str],
    color_by: str,
    granularity: str,
    group_by: str,
    sort_by: str,
    sort_order: str,
    category_order: list[str],
    assignee_order: list[str],
    date_range: list[int],
    show_today_line: list
) -> tuple:
    """フィルター変更時にダッシュボードを更新"""

    # 日付範囲の計算
    start_date = min_date + pd.Timedelta(days=date_range[0])
    end_date = min_date + pd.Timedelta(days=date_range[1])

    # フィルタリング
    filtered_df = df[
        (df["四半期"].isin(quarters or [])) &
        (df["担当者"].isin(assignees or [])) &
        (df["カテゴリ"].isin(categories or [])) &
        (df["開始日"] >= start_date) &
        (df["終了日"] <= end_date)
    ].copy()

    # ソート
    filtered_df = sort_dataframe(
        filtered_df,
        sort_by=sort_by,
        sort_order=sort_order,
        category_order=category_order or DEFAULT_CATEGORY_ORDER,
        assignee_order=assignee_order or DEFAULT_ASSIGNEE_ORDER
    )

    # ガントチャート生成
    show_today = True in (show_today_line or [])
    fig = create_gantt_chart(
        filtered_df,
        color_by=color_by,
        granularity=granularity,
        group_by=group_by,
        show_today_line=show_today
    )

    # サマリー計算
    total = len(filtered_df)

    if not filtered_df.empty:
        date_range_str = f"{filtered_df['開始日'].min().strftime('%Y/%m/%d')} 〜 {filtered_df['終了日'].max().strftime('%Y/%m/%d')}"

        # 担当者別サマリー
        assignee_counts = filtered_df.groupby("担当者").size()
        assignee_badges = [
            dbc.Badge(
                f"{k}: {v}",
                color="primary" if k == "松永" else "danger",
                className="me-1"
            )
            for k, v in assignee_counts.items()
        ]

        # カテゴリ別サマリー
        category_counts = filtered_df.groupby("カテゴリ").size()
        category_badges = [
            dbc.Badge(
                f"{k[:4]}…: {v}" if len(k) > 5 else f"{k}: {v}",
                style={"backgroundColor": CATEGORY_COLORS.get(k, "#999")},
                className="me-1 mb-1"
            )
            for k, v in category_counts.items()
        ]
    else:
        date_range_str = "-"
        assignee_badges = "-"
        category_badges = "-"

    return (
        fig,
        f"{total}",
        date_range_str,
        assignee_badges,
        category_badges
    )


@callback(
    Output("download-excel", "data"),
    Input("download-btn", "n_clicks"),
    [State("quarter-filter", "value"),
     State("assignee-filter", "value"),
     State("category-filter", "value"),
     State("date-range-slider", "value"),
     State("sort-by", "value"),
     State("sort-order", "value"),
     State("category-order", "value"),
     State("assignee-order", "value"),
     State("granularity", "value"),
     State("color-by", "value")],
    prevent_initial_call=True
)
def download_excel(
    n_clicks: int,
    quarters: list[str],
    assignees: list[str],
    categories: list[str],
    date_range: list[int],
    sort_by: str,
    sort_order: str,
    category_order: list[str],
    assignee_order: list[str],
    granularity: str,
    color_by: str
) -> dict:
    """フィルター適用後のデータをExcelガントチャートでダウンロード"""

    # 日付範囲の計算
    start_date = min_date + pd.Timedelta(days=date_range[0])
    end_date = min_date + pd.Timedelta(days=date_range[1])

    filtered_df = df[
        (df["四半期"].isin(quarters or [])) &
        (df["担当者"].isin(assignees or [])) &
        (df["カテゴリ"].isin(categories or [])) &
        (df["開始日"] >= start_date) &
        (df["終了日"] <= end_date)
    ].copy()

    # ソート（画面と同じ順序）
    filtered_df = sort_dataframe(
        filtered_df,
        sort_by=sort_by,
        sort_order=sort_order,
        category_order=category_order or DEFAULT_CATEGORY_ORDER,
        assignee_order=assignee_order or DEFAULT_ASSIGNEE_ORDER
    )

    # Excelガントチャート生成
    excel_data = create_excel_gantt_chart(
        filtered_df,
        granularity=granularity,
        color_by=color_by
    )

    filename = f"eezo_gantt_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return dcc.send_bytes(excel_data.getvalue(), filename)


# =============================================================================
# メイン実行
# =============================================================================

if __name__ == "__main__":
    # 出力ディレクトリの作成
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 60)
    print("  EEZO 2026 タスクダッシュボード")
    print("=" * 60)
    print(f"  データ: {len(df)} タスク")
    print(f"  期間: {df['開始日'].min().strftime('%Y/%m/%d')} 〜 {df['終了日'].max().strftime('%Y/%m/%d')}")
    print(f"  担当者: {', '.join(df['担当者'].unique())}")
    print(f"  カテゴリ: {len(df['カテゴリ'].unique())} 種類")
    print("=" * 60)
    print("  アクセス: http://127.0.0.1:8050")
    print("=" * 60)

    app.run(debug=True, host="0.0.0.0", port=8050)
